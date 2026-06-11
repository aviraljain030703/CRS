"""
PDF and Excel report generation.

Demonstrates: OOP, Inheritance, Polymorphism, File Handling, Generators
"""

import io
import logging
from abc import ABC, abstractmethod
from datetime import datetime

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Abstract base — Abstraction, Inheritance, Polymorphism
# ---------------------------------------------------------------------------

class BaseReportGenerator(ABC):
    """
    Abstract base for report generators.
    Subclasses must implement `generate()`.
    Demonstrates: Abstraction (abstractmethod), Inheritance
    """

    def __init__(self, title: str):
        self._title = title  # Encapsulated attribute
        self._generated_at = datetime.utcnow()

    @property
    def title(self) -> str:
        return self._title

    @abstractmethod
    def generate(self, complaints: list) -> bytes:
        """Generate the report and return raw bytes."""

    def _rows_generator(self, complaints):
        """
        Generator that yields formatted row tuples from complaints.
        Demonstrates: Generator (yield)
        """
        for c in complaints:
            yield (
                c.complaint_number,
                c.user.name if c.user else "N/A",
                c.title,
                c.category,
                c.priority,
                c.status,
                c.created_at.strftime("%Y-%m-%d") if c.created_at else "",
            )


# ---------------------------------------------------------------------------
# PDF Report
# ---------------------------------------------------------------------------

class PDFReportGenerator(BaseReportGenerator):
    """
    Generates a PDF report using ReportLab.
    Demonstrates: Inheritance, Polymorphism (overrides generate())
    """

    def generate(self, complaints: list) -> bytes:
        """Build and return a PDF byte string."""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable,
        )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()

        header_style = ParagraphStyle(
            "Header", parent=styles["Heading1"],
            fontSize=18, textColor=colors.HexColor("#1a237e"), spaceAfter=4
        )
        sub_style = ParagraphStyle(
            "Sub", parent=styles["Normal"],
            fontSize=10, textColor=colors.grey
        )

        story = [
            Paragraph(self._title, header_style),
            Paragraph(f"Generated: {self._generated_at.strftime('%d %b %Y %H:%M UTC')}", sub_style),
            Paragraph(f"Total records: {len(complaints)}", sub_style),
            Spacer(1, 6*mm),
            HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1a237e")),
            Spacer(1, 4*mm),
        ]

        col_headers = [
            "Complaint #", "Student", "Title", "Category", "Priority", "Status", "Date"
        ]
        data = [col_headers]
        for row in self._rows_generator(complaints):   # uses the generator
            data.append(list(row))

        # Colour rows alternately
        row_styles = []
        for i in range(1, len(data)):
            bg = colors.HexColor("#e8eaf6") if i % 2 == 0 else colors.white
            row_styles.append(("BACKGROUND", (0, i), (-1, i), bg))

        table = Table(
            data,
            colWidths=[35*mm, 35*mm, 55*mm, 30*mm, 25*mm, 30*mm, 25*mm],
            repeatRows=1,
        )
        table.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1a237e")),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0), 10),
            ("ALIGN",       (0, 0), (-1, -1), "LEFT"),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("GRID",        (0, 0), (-1, -1), 0.25, colors.HexColor("#c5cae9")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#e8eaf6")]),
            ("FONTSIZE",    (0, 1), (-1, -1), 8),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ] + row_styles))

        story.append(table)
        doc.build(story)
        return buffer.getvalue()


# ---------------------------------------------------------------------------
# Excel Report
# ---------------------------------------------------------------------------

class ExcelReportGenerator(BaseReportGenerator):
    """
    Generates an XLSX report using openpyxl.
    Demonstrates: Inheritance, Polymorphism (overrides generate())
    """

    def generate(self, complaints: list) -> bytes:
        """Build and return an XLSX byte string."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Complaints"

        # Styling helpers
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1A237E")
        alt_fill    = PatternFill("solid", fgColor="E8EAF6")
        thin_border = Border(
            left=Side(style="thin", color="C5CAE9"),
            right=Side(style="thin", color="C5CAE9"),
            top=Side(style="thin", color="C5CAE9"),
            bottom=Side(style="thin", color="C5CAE9"),
        )

        # Title row
        ws.merge_cells("A1:G1")
        ws["A1"] = self._title
        ws["A1"].font = Font(bold=True, size=14, color="1A237E")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Generated: {self._generated_at.strftime('%d %b %Y %H:%M UTC')} | Total: {len(complaints)}"
        ws["A2"].font = Font(italic=True, color="555555")
        ws["A2"].alignment = Alignment(horizontal="center")

        # Column headers
        headers = ["Complaint #", "Student", "Title", "Category", "Priority", "Status", "Date"]
        col_widths = [18, 22, 40, 16, 12, 16, 14]
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = width

        # Data rows — use the generator
        priority_colors = {"Low": "4CAF50", "Medium": "FF9800", "High": "F44336", "Critical": "880E4F"}
        for row_idx, row in enumerate(self._rows_generator(complaints), start=5):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                # Colour priority cell
                if col_idx == 5:
                    hex_c = priority_colors.get(str(value), "FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=hex_c)
                    cell.font = Font(color="FFFFFF", bold=True)

        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:G{ws.max_row}"

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


# ---------------------------------------------------------------------------
# Factory function — Polymorphism (returns correct subclass by format string)
# ---------------------------------------------------------------------------

def get_report_generator(fmt: str, title: str = "Complaint Report") -> BaseReportGenerator:
    """
    Factory: return the right generator for the requested format.
    Demonstrates: Polymorphism, Functions
    """
    generators = {
        "pdf":  PDFReportGenerator,
        "excel": ExcelReportGenerator,
        "xlsx":  ExcelReportGenerator,
    }
    cls = generators.get(fmt.lower())
    if cls is None:
        raise ValueError(f"Unknown report format '{fmt}'. Supported: {list(generators)}")
    return cls(title)
